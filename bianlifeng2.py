# -*-coding:utf8-*-

import openpyxl
from itertools import groupby
from openpyxl.styles import Alignment
from win32com.client import Dispatch
import decimal

timeSet = []
timeList = []
nameSet = set()

# 部门字典
deptDict = {"配料": "烘焙", "搅拌": "烘焙", "成型": "烘焙", "烘烤": "烘焙", "包装": "烘焙", "冷加工": "烘焙"}
# 岗位字典
# postDict = {"王圣平": "", "周彬": "", "苏加强": "", "吴延丽": "", "周兴慧": "素切", "裴勤": "素切",
#             "何俊花": "素切", "胡赛霞": "素切", "邢小芳": "素切", "冉秋英": "素切", "张义琴": "素切", "张增产": "素切",
#             "吴赛英": "素切", "王丽君": "素切", "李晓云": "荤切", "孟志峰": "荤切", "杨叙芬": "荤切", "童永兰": "荤切",
#             "王水琴": "荤切", "李相红": "荤切", "简洪琼": "荤切", "吴亭亭": "荤切", "余朋艳": "蛋品", "廖满青": "蛋品", "张风清": "蛋品", "叶翠兰": "蛋品",
#             "罗玉兰": "蛋品", "黄亚": "蛋品", "李喜各": "蛋品", "徐秀兰": "蛋品", "贺转运": "厨师", "闫宏良": "厨师",
#             "夏盛衡": "厨师", "李永华": "厨师", "胡诗月": "厨师", "李修朱": "厨师", "李海营": "厨师", "刘佳毅": "厨师",
#             "李有生": "厨师", "王志伟": "米饭", "张洪立": "米饭", "董亚伟": "米饭", "罗顺林": "米饭", "王勤": "拉菜冷菜",
#             "张巧兰": "拉菜冷菜", "张文红": "拉菜冷菜", "吴六四": "拉菜冷菜", "董变芬": "拉菜冷菜", "李慧芳": "拉菜冷菜", "李瑞平": "煮面",
#             "盛参": "煮面", "周宝荣": "洗框", "汪秀兰": "洗框", "张莉娜": "组装", "位燕梅": "组装", "蔡雁": "组装", "张秀平": "组装",
#             "王克英": "组装", "李祥锋": "组装", "张娜": "组装", "姚邦云": "组装", "周妮": "组装", "张小蜜": "组装", "陈波": "组装", "王培云": "组装",
#             "栾婷婷": "组装", "张玲丽": "组装", "周雪梅": "组装", "王玉": "组装", "张桂芳": "组装", "徐宏英": "组装", "隋庆全": "组装", "林青花": "组装",
#             "王乐侠": "组装", "朱正英": "组装", "王美香": "组装", "汤恒菊": "组装", "李保真": "组装", "谢治惠": "组装", "王金格": "组装", "詹金涛": "组装",
#             }

# 工号字典
# numDict = {"张秀琴": "A0006", "唐申申": "S0129", "向锋": "X0126", "甘高平": "G0041", "王海鑫": "S0134", "许波": "Y0169",
#             "毛勤": "A0284", "曹然然": "Y0161", "李井云": "S0097", "滕娟娟": "S0111", "刘小丽": "G0050", "余能勤": "H0003",
#             "赵迪": "G0038", "王柒金": "A0310", "叶徐林": "S0148", "曹荣基": "X0127", "任俊荣": "X0121", "徐希侠": "G0054",
#             "刘年": "X0069", "沈金金": "X0089", "李亚南": "Y0206", "彭国清": "X0128", "刘士群": "G0049", "朱庆生": "Y0205", "张玲玲": "X0111", "张珊": "B0025",
#             "刘艳": "X0078", "谢振芝": "S0110", "盛魏敏": "G0055", "李娟娟": "S0124", "张娜": "S0122", "杨进英": "S0123",
#             "朱金梅": "S0149", "高月月": "A0309", "冉隆琴": "Y0218", "翟翠萍": "S0128", "王文雷": "X0106", "王雪雪": "B0011",
#             "郭志勇": "A0290", "牛园园": "X0083", "花丽玉": "X0088", "王晶晶": "S0101", "刘娟": "X0070", "马云": "Y0177",
#             "俞建平": "A0282", "张伟": "A0287", "张康宝": "A0281",
#            "李亚男": "Y0206"
#             }
postDict = {}

sortDict = {
    "": 0,
    "素切": 1,
    "荤切": 2,
    "蛋品": 3,
    "厨师": 4,
    "米饭": 5,
    "拉菜冷菜": 6,
    "煮面": 7,
    "洗框": 8,
    "组装": 9,
    "保洁": 10,
    "仓库": 11,
    "配料": 12,
    "搅拌": 13
}

file_name = 'D:/example4.xlsx'

if __name__ == "__main__":
    xlApp = Dispatch("Excel.Application")
    xlApp.Visible = False
    xlBook = xlApp.Workbooks.Open(file_name)
    xlBook.Save()
    xlBook.Close()
    wb = openpyxl.load_workbook(file_name, data_only=True)
    # sheets = wb.sheetnames
    # print(sheets, type(sheets))
    for sheet in wb:
        if not sheet.title.__contains__("."):
            print(sheet.title)
            continue
        else:
            timeSet.append(sheet.title)
            postName = ''
            for name in sheet['B']:
                if name.value is None or name.value.__contains__("姓名") or name.value.__contains__("2021"):
                    pass
                else:
                    personTime = {}
                    personTime['day'] = sheet.title
                    if not (sheet['A'+str(name.row)].value is None or sheet['A'+str(name.row)].value == ''):
                        postName = sheet['A'+str(name.row)].value
                    if postName == '保洁' or postName == '仓库':
                        continue
                    if name.value.strip() not in postDict:
                        postDict[name.value] = postName
                    personTime['post'] = postDict[name.value.strip()]
                    personTime['dept'] = '鲜食'
                    personTime['name'] = name.value.strip()
                    personTime['num'] = ''
                    personTime['order'] = sortDict[postDict[name.value]]
                    if sheet['I'+str(name.row)].value is None or int(sheet['I'+str(name.row)].value) == 0:
                        personTime['time'] = sheet['G'+str(name.row)].value
                    else:
                        # personTime['time'] = int(sheet['I'+str(name.row)].value)/60
                        personTime['time'] = decimal.Decimal(sheet['I'+str(name.row)].value)/decimal.Decimal(60)
                    personTime['night'] = sheet['H'+str(name.row)].value
                    timeList.append(personTime)

                    # print(personTime)
    # 多字段分组
    user_sort = sorted(timeList, key=lambda x: (x['order'], x["name"]))
    # 多字段分组
    name_group = groupby(user_sort, key=lambda x: (x["name"]))
    # for key, group in name_group:
    #     print(key, list(group))

    #输出excel
    # from openpyxl import Workbook
    # wb = Workbook()
    # ws = wb.active

    # ws = wb.create_sheet('总表2')
    ws = wb['总表']
    title = ['序号', '工号', '姓名', '部门', '岗位', '入职日期', '离职日期', '工时类别', '工时夜班']
    title.extend(timeSet)
    print('总行数' + str(ws.max_row))
    maxRow = ws.max_row
    # ws.append(title)
    for key, group in name_group:
        nameTimeList = list(group)
        nameSet.add(key)
        rowList = []
        nightList = []
        rowList.extend(['', nameTimeList[0]['num'], key, nameTimeList[0]['dept'], nameTimeList[0]['post'], '', '', '综合工时制', '工时'])
        nightList.extend(['', '', '', '', '', '', '', '', '夜班'])
        for time in timeSet:
            bl = True
            for row in nameTimeList:
                if row['day'] == time:
                    bl = False
                    rowList.append(row['time'])
                    nightList.append(row['night'])
                    break
            if bl:
                rowList.append('')
                nightList.append('')
        ws.append(rowList)
        ws.append(nightList)
    num = 1
    alignment_center = Alignment(horizontal='center', vertical='center')
    print('总人数={}'.format(len(nameSet)))
    cellList = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H']

    for key in nameSet:
        for c in cellList:
            ws.merge_cells((c+'{}:'+c+'{}').format(2*num+maxRow, 2*num+maxRow+1))
        num = num + 1

    # for row in ws.iter_rows():
    #     for cell in row:
    #         cell.alignment = alignment_center
    # wb.save('D:/汇总数据3.xlsx')
    wb.save(file_name)



