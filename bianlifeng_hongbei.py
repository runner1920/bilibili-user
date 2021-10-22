# -*-coding:utf8-*-

import openpyxl
from itertools import groupby
from openpyxl.styles import Alignment
from win32com.client import Dispatch
import decimal

timeSet = []
timeList = []
nameSet = set()

# 部门字典
deptDict = {"配料": "烘焙", "搅拌": "烘焙", "成型": "烘焙", "烘烤": "烘焙", "包装": "烘焙", "冷加工": "烘焙"}

postDict = {}

sortDict = {
    "": 0,
    "素切": 1,
    "荤切": 2,
    "蛋品": 3,
    "厨师": 4,
    "米饭": 5,
    "拉菜冷菜": 6,
    "煮面": 7,
    "洗框": 8,
    "组装": 9,
    "保洁": 10,
    "仓库": 11,
    "配料": 12,
    "搅拌": 13
}

file_name = 'D:/example4.xlsx'

if __name__ == "__main__":
    xlApp = Dispatch("Excel.Application")
    xlApp.Visible = False
    xlBook = xlApp.Workbooks.Open(file_name)
    xlBook.Save()
    xlBook.Close()
    wb = openpyxl.load_workbook(file_name, data_only=True)
    # sheets = wb.sheetnames
    # print(sheets, type(sheets))
    for sheet in wb:
        if not sheet.title.__contains__("0"):
            print(sheet.title)
            continue
        else:
            timeSet.append(sheet.title)
            postName = ''
            for name in sheet['C']:
                if str(name.value) is None or str(name.value).__contains__("姓名") or str(name.value).__contains__("2021"):
                    pass
                else:
                    personTime = {}
                    # personTime['day'] = sheet.title
                    # if not (sheet['A'+str(name.row)].value is None or sheet['A'+str(name.row)].value == ''):
                    #     postName = sheet['A'+str(name.row)].value
                    # if postName == '保洁' or postName == '仓库':
                    #     continue
                    # if name.value.strip() not in postDict:
                    #     postDict[name.value] = postName
                    # personTime['post'] = postDict[name.value.strip()]
                    # personTime['dept'] = '鲜食'
                    personTime['name'] = name.value
                    # personTime['num'] = ''
                    # personTime['order'] = sortDict[postDict[name.value]]
                    # if sheet['I'+str(name.row)].value is None or int(sheet['I'+str(name.row)].value) == 0:
                    #     personTime['time'] = sheet['G'+str(name.row)].value
                    # else:
                    #     # personTime['time'] = int(sheet['I'+str(name.row)].value)/60
                    #     personTime['time'] = decimal.Decimal(sheet['I'+str(name.row)].value)/decimal.Decimal(60)
                    # personTime['night'] = sheet['H'+str(name.row)].value
                    timeList.append(personTime)
                    nameSet.add(name.value)

                    # print(personTime)
    # 多字段分组
    # user_sort = sorted(timeList, key=lambda x: (x['order'], x["name"]))
    # # 多字段分组
    # name_group = groupby(user_sort, key=lambda x: (x["name"]))
    # for key, group in name_group:
    #     print(key, list(group))

    #输出excel
    # from openpyxl import Workbook
    # wb = Workbook()
    # ws = wb.active
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    title = ['姓名']
    ws.append(title)
    for row in nameSet:
        rowList = []
        rowList.append(row)
        ws.append(rowList)
    wb.save('D:/汇总数据.xlsx')




