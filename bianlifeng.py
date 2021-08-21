# -*-coding:utf8-*-

import openpyxl
from itertools import groupby
from openpyxl.styles import Alignment

timeSet = []
timeList = []
nameSet = set()

# 部门字典
deptDict = {"配料": "烘焙", "搅拌": "烘焙", "成型": "烘焙", "烘烤": "烘焙", "包装": "烘焙", "冷加工": "烘焙"}
# 岗位字典
postDict = {"张秀琴": "配料", "唐申申": "配料", "向锋": "配料", "甘高平": "搅拌", "王海鑫": "搅拌", "许波": "搅拌",
            "毛勤": "成型", "曹然然": "成型", "李井云": "成型", "滕娟娟": "成型", "刘小丽": "成型", "余能勤": "成型",
            "赵迪": "烘烤", "王柒金": "烘烤", "叶徐林": "烘烤", "曹荣基": "烘烤", "任俊荣": "包装", "徐希侠": "包装",
            "刘年": "包装", "沈金金": "包装", "李亚南": "包装", "彭国清": "包装", "刘士群": "包装", "朱庆生": "包装", "张玲玲": "包装", "张珊": "包装",
            "刘艳": "冷加工", "谢振芝": "冷加工", "盛魏敏": "冷加工", "李娟娟": "冷加工", "张娜": "冷加工", "杨进英": "冷加工",
            "朱金梅": "冷加工", "高月月": "冷加工", "冉隆琴": "冷加工", "翟翠萍": "冷加工", "王文雷": "冷加工", "王雪雪": "冷加工",
            "郭志勇": "蛋糕", "牛园园": "蛋糕", "花丽玉": "蛋糕", "王晶晶": "蛋糕", "刘娟": "蛋糕", "马云": "蛋糕",
            "俞建平": "面包主管", "张伟": "蛋糕主管", "张康宝": "面包领班",
            "李亚男": "包装"
            }

# 工号字典
numDict = {"张秀琴": "A0006", "唐申申": "S0129", "向锋": "X0126", "甘高平": "G0041", "王海鑫": "S0134", "许波": "Y0169",
            "毛勤": "A0284", "曹然然": "Y0161", "李井云": "S0097", "滕娟娟": "S0111", "刘小丽": "G0050", "余能勤": "H0003",
            "赵迪": "G0038", "王柒金": "A0310", "叶徐林": "S0148", "曹荣基": "X0127", "任俊荣": "X0121", "徐希侠": "G0054",
            "刘年": "X0069", "沈金金": "X0089", "李亚南": "Y0206", "彭国清": "X0128", "刘士群": "G0049", "朱庆生": "Y0205", "张玲玲": "X0111", "张珊": "B0025",
            "刘艳": "X0078", "谢振芝": "S0110", "盛魏敏": "G0055", "李娟娟": "S0124", "张娜": "S0122", "杨进英": "S0123",
            "朱金梅": "S0149", "高月月": "A0309", "冉隆琴": "Y0218", "翟翠萍": "S0128", "王文雷": "X0106", "王雪雪": "B0011",
            "郭志勇": "A0290", "牛园园": "X0083", "花丽玉": "X0088", "王晶晶": "S0101", "刘娟": "X0070", "马云": "Y0177",
            "俞建平": "A0282", "张伟": "A0287", "张康宝": "A0281",
           "李亚男": "Y0206"
            }

if __name__ == "__main__":
    wb = openpyxl.load_workbook('D:/example.xlsx')
    # sheets = wb.sheetnames
    # print(sheets, type(sheets))
    for sheet in wb:
        if sheet.title.__contains__("Sheet"):
            print(sheet.title)
            continue
        else:
            timeSet.append(sheet.title)
            for name in sheet['B']:
                if name.value is None or name.value.__contains__("姓名") or name.value.__contains__("2021"):
                    pass
                else:
                    personTime = {}
                    personTime['day'] = sheet.title
                    personTime['post'] = postDict[name.value]
                    personTime['dept'] = '烘焙'
                    personTime['num'] = numDict[name.value]
                    personTime['name'] = name.value
                    personTime['time'] = sheet['F'+str(name.row)].value
                    personTime['night'] = sheet['G'+str(name.row)].value
                    timeList.append(personTime)

                    # print(personTime)
    # 多字段分组
    user_sort = sorted(timeList, key=lambda x: (x['post'], x["name"]))
    # 多字段分组
    name_group = groupby(user_sort, key=lambda x: (x["name"]))
    # for key, group in name_group:
    #     print(key, list(group))

    #输出excel
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    title = ['序号', '工号', '姓名', '部门', '岗位', '入职日期', '离职日期', '工时类别', '工时夜班']
    title.extend(timeSet)
    ws.append(title)
    for key, group in name_group:
        nameTimeList = list(group)
        nameSet.add(key)
        rowList = []
        nightList = []
        rowList.extend(['', nameTimeList[0]['num'], key, nameTimeList[0]['dept'], nameTimeList[0]['post'], '', '', '综合工时制', '工时'])
        nightList.extend(['', '', '', '', '', '', '', '', '夜班'])
        for time in timeSet:
            bl = True
            for row in nameTimeList:
                if row['day'] == time:
                    bl = False
                    rowList.append(row['time'])
                    nightList.append(row['night'])
                    break
            if bl:
                rowList.append('')
                nightList.append('')
        ws.append(rowList)
        ws.append(nightList)
    num = 1
    alignment_center = Alignment(horizontal='center', vertical='center')
    for key in nameSet:
        ws.merge_cells('A{}:A{}'.format(2*num, 2*num+1))
        ws.merge_cells('B{}:B{}'.format(2*num, 2*num+1))
        ws.merge_cells('C{}:C{}'.format(2*num, 2*num+1))
        ws.merge_cells('D{}:D{}'.format(2*num, 2*num+1))
        ws.merge_cells('E{}:E{}'.format(2*num, 2*num+1))
        ws.merge_cells('F{}:F{}'.format(2*num, 2*num+1))
        ws.merge_cells('G{}:G{}'.format(2*num, 2*num+1))
        ws.merge_cells('H{}:H{}'.format(2*num, 2*num+1))
        num = num + 1

    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = alignment_center
    wb.save('D:/汇总数据.xlsx')


